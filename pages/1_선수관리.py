"""
선수 관리 페이지
- 대회별 선수 등록 / 수정 / 삭제
- 와일드카드 설정
"""
import streamlit as st
import db

st.set_page_config(page_title="선수 관리", page_icon="🎾")
st.title("선수 관리")

# ── 사이드바: 대회 선택 ───────────────────────────────────────────────────────
tournaments = db.get_tournaments()
if not tournaments:
    st.warning("먼저 홈에서 대회를 만들어 주세요.")
    st.stop()

t_names = [t["name"] for t in tournaments]
selected_name = st.sidebar.selectbox("대회 선택", t_names)
tournament = next(t for t in tournaments if t["name"] == selected_name)
tid = tournament["id"]

st.sidebar.markdown(f"**날짜:** {tournament.get('date') or '미설정'}")
st.sidebar.markdown(f"**상태:** {'✅ 완료' if tournament['is_finished'] else '🔄 진행 중'}")

# ── 현재 선수 목록 ────────────────────────────────────────────────────────────
st.subheader(f"{selected_name} 선수 목록")
players = db.get_players(tid)

if players:
    for p in players:
        col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
        with col1:
            st.write(f"**{p['name']}**" + (" 🃏" if p["is_wildcard"] else ""))
        with col2:
            st.write(p.get("title") or "-")
        with col3:
            # 수정 버튼
            if st.button("수정", key=f"edit_{p['id']}"):
                st.session_state["editing_player"] = p
        with col4:
            if st.button("삭제", key=f"del_{p['id']}"):
                db.delete_player(p["id"])
                st.rerun()
else:
    st.info("등록된 선수가 없습니다.")

st.divider()

# ── 선수 추가 / 수정 폼 ───────────────────────────────────────────────────────
editing = st.session_state.get("editing_player")
form_title = "선수 수정" if editing else "선수 추가"
st.subheader(form_title)

with st.form("player_form", clear_on_submit=True):
    name = st.text_input("이름", value=editing["name"] if editing else "")
    title = st.text_input("직함 (예: 시드1, 일반2, WC1)", value=editing.get("title", "") if editing else "")
    is_wc = st.checkbox("와일드카드 (보너스 점수 대상)", value=editing["is_wildcard"] if editing else False)

    submitted = st.form_submit_button("저장")
    if submitted:
        if not name.strip():
            st.error("이름을 입력해 주세요.")
        else:
            db.upsert_player(
                tournament_id=tid,
                name=name.strip(),
                title=title.strip(),
                is_wildcard=is_wc,
                player_id=editing["id"] if editing else None,
            )
            st.session_state.pop("editing_player", None)
            st.success("저장했습니다.")
            st.rerun()

if editing and st.button("취소"):
    st.session_state.pop("editing_player", None)
    st.rerun()

# ── 엑셀 데이터 가져오기 ──────────────────────────────────────────────────────
st.divider()
with st.expander("기존 엑셀 파일에서 선수 가져오기"):
    st.caption("모비딕테니스_...xlsx 파일의 '설정' 시트에서 선수 목록을 자동으로 읽어옵니다.")
    uploaded = st.file_uploader("엑셀 파일 업로드", type=["xlsx"])
    if uploaded:
        import openpyxl
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        if "설정" not in wb.sheetnames:
            st.error("'설정' 시트를 찾을 수 없습니다.")
        else:
            ws = wb["설정"]
            imported = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                title_val, id_val, name_val, *_ = row
                if not name_val:
                    continue
                is_wc_val = str(id_val or "").startswith("W")
                imported.append({"title": title_val, "name": name_val, "is_wildcard": is_wc_val})

            st.write(f"{len(imported)}명 발견:")
            for p in imported:
                st.write(f"- {p['name']} ({p['title']}) {'🃏 WC' if p['is_wildcard'] else ''}")

            if st.button("이 선수들을 현재 대회에 추가"):
                for p in imported:
                    db.upsert_player(tid, p["name"], p["title"], p["is_wildcard"])
                st.success(f"{len(imported)}명 추가 완료!")
                st.rerun()
